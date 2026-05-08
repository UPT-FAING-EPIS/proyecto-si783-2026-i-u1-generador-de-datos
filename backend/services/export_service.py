import json
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict, Any

def export_to_json(data: Dict[str, List[Dict[str, Any]]]) -> bytes:
    return json.dumps(data, indent=2, ensure_ascii=False, default=str).encode("utf-8")

def export_to_csv(data: Dict[str, List[Dict[str, Any]]]) -> bytes:
    output = io.StringIO()
    for table_name, rows in data.items():
        if not rows:
            continue
        output.write(f"# Tabla: {table_name}\n")
        writer = csv.DictWriter(output, fieldnames=rows[0].keys(), lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
        output.write("\n")
    return output.getvalue().encode("utf-8")

def export_to_excel(data: Dict[str, List[Dict[str, Any]]]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

    for table_name, rows in data.items():
        if not rows:
            continue
        ws = wb.create_sheet(title=table_name[:31])
        headers = list(rows[0].keys())

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(rows, 2):
            for col_idx, key in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(row.get(key, "")) if row.get(key) is not None else "")
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        for col_idx, header in enumerate(headers, 1):
            max_len = max(len(str(header)), max((len(str(row.get(header, "") or "")) for row in rows), default=0))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def export_to_sql(data: Dict[str, List[Dict[str, Any]]], engine: str = "mysql") -> bytes:
    output = []
    output.append(f"-- SmartGen Export")
    output.append(f"-- Engine: {engine}")
    output.append(f"-- Generated tables: {', '.join(data.keys())}")
    output.append("")

    for table_name, rows in data.items():
        if not rows:
            continue
        output.append(f"-- Tabla: {table_name}")
        output.append(f"-- {len(rows)} registros")
        output.append("")

        for row in rows:
            cols = ", ".join(f"`{k}`" if engine == "mysql" else f'"{k}"' for k in row.keys())
            vals = []
            for v in row.values():
                if v is None:
                    vals.append("NULL")
                elif isinstance(v, bool):
                    vals.append("TRUE" if v else "FALSE")
                elif isinstance(v, (int, float)):
                    vals.append(str(v))
                else:
                    escaped = str(v).replace("'", "''")
                    vals.append(f"'{escaped}'")
            values = ", ".join(vals)
            if engine == "mysql":
                output.append(f"INSERT INTO `{table_name}` ({cols}) VALUES ({values});")
            else:
                output.append(f'INSERT INTO "{table_name}" ({cols}) VALUES ({values});')

        output.append("")

    return "\n".join(output).encode("utf-8")

def export_to_xml(data: Dict[str, List[Dict[str, Any]]]) -> bytes:
    lines = ['<?xml version="1.0" encoding="UTF-8"?>', "<smartgen_export>"]
    for table_name, rows in data.items():
        if not rows:
            continue
        lines.append(f"  <table name=\"{table_name}\">")
        for row in rows:
            lines.append("    <record>")
            for k, v in row.items():
                safe_k = str(k).replace(" ", "_")
                safe_v = str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;") if v is not None else ""
                lines.append(f"      <{safe_k}>{safe_v}</{safe_k}>")
            lines.append("    </record>")
        lines.append("  </table>")
    lines.append("</smartgen_export>")
    return "\n".join(lines).encode("utf-8")
